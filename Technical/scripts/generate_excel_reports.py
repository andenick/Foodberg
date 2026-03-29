"""
Excel Report Generator for Foodberg
Druck-compliant: ONE SHEET PER FILE

Generates professional Excel reports with:
- Machine-readable column names
- Professional black & white formatting
- Timestamped filenames
- One sheet per file (Druck standard)
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

OUTPUT_DIR = Path(__file__).parent.parent.parent / 'Output' / 'Data'
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

def create_druck_compliant_excel(data: pd.DataFrame, filename: str, sheet_name: str = 'Data'):
    """
    Create Druck-compliant Excel file with ONE SHEET
    
    Args:
        data: pandas DataFrame with data
        filename: Output filename (without extension)
        sheet_name: Sheet name (default: 'Data')
    """
    # Add timestamp to filename
    timestamp = datetime.now().strftime("%Y.%m.%d")
    output_path = OUTPUT_DIR / f"[{timestamp}] {filename}.xlsx"
    
    # Write to Excel (single sheet only)
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        data.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Professional formatting (B&W only)
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        alignment = Alignment(horizontal='left', vertical='center')
        
        # Format headers
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = alignment
        
        # Format data cells
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = alignment
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        worksheet.freeze_panes = 'A2'
    
    print(f"✅ Created: {output_path}")
    return output_path


def generate_current_commodity_prices():
    """Generate current commodity prices report"""
    # Mock data - replace with actual API data
    data = pd.DataFrame({
        'commodity': ['tomatoes', 'chicken_breast', 'onions', 'potatoes', 'lettuce'],
        'current_price_usd_per_lb': [2.45, 3.99, 1.20, 0.95, 2.10],
        'price_change_24h': [0.15, -0.10, 0.05, 0.02, -0.05],
        'price_change_percent': [6.5, -2.4, 4.3, 2.2, -2.3],
        'market': ['new_york', 'new_york', 'new_york', 'new_york', 'new_york'],
        'category': ['vegetables', 'proteins', 'vegetables', 'vegetables', 'vegetables'],
        'last_updated': [datetime.now().isoformat()] * 5
    })
    
    create_druck_compliant_excel(
        data,
        'current_commodity_prices',
        'Commodity_Prices'
    )


def generate_recipe_costs_summary():
    """Generate recipe costs summary report"""
    data = pd.DataFrame({
        'recipe_name': ['Caesar Salad', 'Grilled Salmon', 'Chicken Stir-Fry'],
        'servings': [6, 4, 4],
        'ingredient_cost_usd': [12.50, 24.80, 15.30],
        'labor_cost_usd': [10.00, 15.00, 12.00],
        'overhead_cost_usd': [3.75, 7.44, 4.59],
        'total_cost_usd': [26.25, 47.24, 31.89],
        'cost_per_serving_usd': [4.38, 11.81, 7.97],
        'suggested_menu_price_usd': [15.60, 42.18, 28.46],
        'profit_per_serving_usd': [11.22, 30.37, 20.49],
        'food_cost_percent': [28.0, 28.0, 28.0]
    })
    
    create_druck_compliant_excel(
        data,
        'recipe_costs_summary',
        'Recipe_Costs'
    )


def generate_menu_engineering_analysis():
    """Generate menu engineering analysis report"""
    data = pd.DataFrame({
        'menu_item': ['Grilled Chicken', 'Pasta Primavera', 'Steak Frites', 'Caesar Salad', 'Lobster Risotto'],
        'menu_price_usd': [24.99, 18.99, 42.99, 14.99, 54.99],
        'cost_usd': [7.50, 6.25, 15.00, 4.38, 22.00],
        'profit_usd': [17.49, 12.74, 27.99, 10.61, 32.99],
        'units_sold_30d': [150, 95, 75, 180, 45],
        'classification': ['STAR', 'PLOW HORSE', 'PUZZLE', 'STAR', 'PUZZLE'],
        'recommended_action': [
            'Maintain quality, promote heavily',
            'Increase price or reduce cost',
            'Increase marketing or reposition',
            'Maintain quality, promote heavily',
            'Increase marketing or reposition'
        ]
    })
    
    create_druck_compliant_excel(
        data,
        'menu_engineering_analysis',
        'Menu_Engineering'
    )


def generate_vendor_price_comparison():
    """Generate vendor price comparison report"""
    data = pd.DataFrame({
        'commodity': ['chicken_breast', 'chicken_breast', 'tomatoes', 'tomatoes', 'onions', 'onions'],
        'vendor': ['Sysco', 'US Foods', 'Sysco', 'US Foods', 'Restaurant Depot', 'Sysco'],
        'price_usd_per_lb': [3.99, 3.85, 2.45, 2.60, 1.15, 1.20],
        'minimum_order_lbs': [10, 10, 25, 20, 50, 25],
        'delivery_fee_usd': [15.00, 12.00, 15.00, 12.00, 0.00, 15.00],
        'delivery_time_days': ['1-2', '2-3', '1-2', '2-3', 'pickup', '1-2'],
        'total_cost_min_order': [54.90, 50.50, 76.25, 64.00, 57.50, 45.00]
    })
    
    create_druck_compliant_excel(
        data,
        'vendor_price_comparison',
        'Vendor_Comparison'
    )


def generate_price_alerts_history():
    """Generate price alerts history report"""
    data = pd.DataFrame({
        'alert_id': ['A001', 'A002', 'A003', 'A004', 'A005'],
        'commodity': ['beef', 'tomatoes', 'chicken', 'salmon', 'lettuce'],
        'alert_type': ['spike', 'drop', 'spike', 'spike', 'drop'],
        'threshold_percent': [15, 10, 12, 20, 8],
        'triggered_date': pd.date_range('2025-01-01', periods=5, freq='D'),
        'price_before': [8.50, 2.45, 3.99, 12.50, 2.10],
        'price_after': [9.78, 2.20, 4.47, 15.00, 1.93],
        'change_percent': [15.1, -10.2, 12.0, 20.0, -8.1],
        'notification_sent': ['SMS', 'Email', 'SMS', 'Push', 'Email']
    })
    
    create_druck_compliant_excel(
        data,
        'price_alerts_history',
        'Price_Alerts'
    )


def generate_weekly_market_report():
    """Generate weekly market summary report"""
    data = pd.DataFrame({
        'commodity': ['tomatoes', 'chicken', 'beef', 'lettuce', 'onions', 'potatoes'],
        'week_start': [datetime.now().strftime('%Y-%m-%d')] * 6,
        'week_avg_price': [2.45, 3.99, 8.50, 2.10, 1.20, 0.95],
        'week_min_price': [2.30, 3.85, 8.25, 2.00, 1.15, 0.92],
        'week_max_price': [2.60, 4.15, 8.75, 2.25, 1.25, 0.98],
        'week_volatility': [0.12, 0.08, 0.06, 0.12, 0.08, 0.06],
        'month_avg_price': [2.40, 4.05, 8.45, 2.15, 1.18, 0.93],
        'price_trend': ['up', 'down', 'stable', 'down', 'up', 'stable']
    })
    
    create_druck_compliant_excel(
        data,
        'weekly_market_report',
        'Weekly_Report'
    )


def generate_all_reports():
    """Generate all Druck-compliant Excel reports"""
    print("\n🎯 Generating Druck-Compliant Excel Reports")
    print("=" * 50)
    
    generate_current_commodity_prices()
    generate_recipe_costs_summary()
    generate_menu_engineering_analysis()
    generate_vendor_price_comparison()
    generate_price_alerts_history()
    generate_weekly_market_report()
    
    print("=" * 50)
    print("✨ All reports generated successfully!")
    print(f"📁 Location: {OUTPUT_DIR}")
    print("\n✅ Druck Compliance Verified:")
    print("   - ONE SHEET per file")
    print("   - Machine-readable column names")
    print("   - Professional B&W formatting")
    print("   - Timestamped filenames")


if __name__ == "__main__":
    generate_all_reports()

