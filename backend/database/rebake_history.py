#!/usr/bin/env python3
"""
Foodberg DB rebake — maximal food-data coverage (2026-06-12 campaign).

Run LOCALLY (not in the container) against backend/data/foodberg.db, then
rebuild/redeploy the backend image. Raw sqlite3 for speed (~2M rows).

What it loads (all from Robin's canonical store, all public-domain/CC-BY):

  1. USDA NASS HISTORY  (DATA/USDA_NASS_HISTORY/nass_hist_*.json)
       -> wasde_data    PRICE RECEIVED / PRODUCTION / YIELD series,
                        NATIONAL 1908-2026 + STATE 1950-2026 (top-15).
                        Replaces the same-category rows from the single-year
                        snapshot (which remain for all OTHER categories).
  2. FAOSTAT producer prices  (Prices_E_All_Data_(Normalized).csv)
       -> global_prices  Element 'Producer Price (USD/tonne)', annual value,
                         per country x item. source='FAOSTAT'.
  3. FAOSTAT consumer food price indices  (ConsumerPriceIndices CSV)
       -> global_prices  Item 23013 'Consumer Prices, Food Indices (2015=100)',
                         monthly per country. source='FAOSTAT CPI'.
  4. World Bank Pink Sheet  (CMO-Historical-Data-Monthly.xlsx)
       -> global_prices  every monthly commodity series + indices,
                         region='World'. source='World Bank Pink Sheet'.
  5. BLS Average Prices  (DATA/BLS_AP/ap_fred_*.json)
       -> retail_prices  ~50 retail food items, monthly. `location` is the
                         artifact's own `area` (U.S. city average for the
                         national items; the Census-region name for the four
                         APU0[1-4]00712311 tomato series). source='BLS AP'.
  6. FRED + BLS macro / food CPI + PPI  (backend/data/collected/*.json)
       -> economic_indicators   upserted on (series_id, date, source).
                         Four artifacts: the 2025-12 fred_data/bls_data
                         snapshots plus fred_food_extra/bls_food_extra written
                         by Technical/scripts/ingest_bls_monthly_series.py.
  7. Composite indices  (derived from 2/4 above + economic_indicators)
       -> composite_indices     recomputed and upserted every rebake.

USDA PSD (quantities) stays in Robin's raw store — acquired but not baked
(no current page consumes supply/demand quantities; honest deferral).

No synthetic values: every row is copied verbatim from the source artifact.

2026-07-24 (P0-5): steps 6 and 7 were added. Previously `economic_indicators`
and `composite_indices` were only COUNTED in the final census and never
refreshed, so `composite_indices.computed_at` was frozen at 2026-03-28 and the
headline Food Price Index page fell behind the Explorer. Both are now refreshed
in-transaction, in dependency order, on every rebake.
"""

import csv
import datetime
import json
import os
import sqlite3
import sys
from pathlib import Path

BACKEND = Path(__file__).resolve().parent.parent
if str(BACKEND) not in sys.path:          # so `indices.composite` resolves
    sys.path.insert(0, str(BACKEND))

DB_PATH = BACKEND / "data" / "foodberg.db"

# Robin's canonical data store. Never hardcode an absolute workspace path here:
# this file ships in the public repo. Set ROBIN_DATA_PATH, or rely on the
# relative default when the project sits inside the workspace.
_robin_env = os.environ.get("ROBIN_DATA_PATH", "").strip()
ROBIN = Path(_robin_env) if _robin_env else (
    BACKEND.parent.parent.parent / "Council" / "Robin" / "DATA")
if not ROBIN.is_dir():
    raise FileNotFoundError(
        "ROBIN_DATA_PATH not set and default not found. "
        "Set ROBIN_DATA_PATH env var to the Robin DATA directory."
    )

COLLECTED = BACKEND / "data" / "collected"
COLLECTED_FRED = COLLECTED / "fred_data.json"
COLLECTED_BLS = COLLECTED / "bls_data.json"
# 2026-07-24: monthly BLS series added by
# Technical/scripts/ingest_bls_monthly_series.py (keyless FRED mirror for
# WPU01130217; keyless BLS public API v1 for CUUR0000SEFV01/02, which FRED does
# not mirror). Separate artifacts so the 2025-12 snapshots above keep their own
# retrieval provenance. Each (path, source) pair names the endpoint the data
# actually came from.
COLLECTED_FRED_EXTRA = COLLECTED / "fred_food_extra.json"
COLLECTED_BLS_EXTRA = COLLECTED / "bls_food_extra.json"
COLLECTED_SOURCES = (
    (COLLECTED_FRED, "FRED"),
    (COLLECTED_BLS, "BLS"),
    (COLLECTED_FRED_EXTRA, "FRED"),
    (COLLECTED_BLS_EXTRA, "BLS"),
)

NASS_DIR = ROBIN / "USDA_NASS_HISTORY"
FAO_PRICES_CSV = ROBIN / "FAO/FAOSTAT_BULK/Prices_E_All_Data_(Normalized).csv"
FAO_CPI_CSV = ROBIN / "FAO/FAOSTAT_BULK/ConsumerPriceIndices_E_All_Data_(Normalized).csv"
PINKSHEET_XLSX = ROBIN / "WORLD_BANK_PINKSHEET/CMO-Historical-Data-Monthly.xlsx"
BLS_AP_DIR = ROBIN / "BLS_AP"

NOW = datetime.datetime.now().isoformat(sep=" ", timespec="seconds")
BATCH = 5000

MONTHS = {m: i for i, m in enumerate(
    ["January", "February", "March", "April", "May", "June", "July",
     "August", "September", "October", "November", "December"], start=1)}


def log(msg: str) -> None:
    print(f"[rebake] {msg}", flush=True)


def numeric(value_str):
    if value_str is None:
        return None
    s = str(value_str).replace(",", "").strip()
    if s in ("(D)", "(S)", "(NA)", "", "None", "(Z)", "(X)"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# 1. NASS history -> wasde_data
# ---------------------------------------------------------------------------

def import_nass(con: sqlite3.Connection) -> int:
    files = sorted(NASS_DIR.glob("nass_hist_*.json"))
    if not files:
        log("NASS: no files found — SKIPPING")
        return 0
    cur = con.cursor()
    total = 0
    cols = ("commodity, statistic_category, value, numeric_value, unit, location, "
            "state_code, agg_level, year, reference_period, short_desc, source_desc, "
            "sector, group_desc, class_desc, freq_desc, load_time, imported_at, "
            "prodn_practice, util_practice, domain_desc, domaincat_desc")
    ins = f"INSERT INTO wasde_data ({cols}) VALUES ({','.join(['?'] * 22)})"

    for fp in files:
        payload = json.loads(fp.read_text(encoding="utf-8"))
        commodity = payload["commodity"].upper()
        cats = tuple(payload["categories"])
        # Replace the time-series categories at NATIONAL/STATE level; the 2025
        # snapshot keeps everything else (county detail, other categories).
        cur.execute(
            f"DELETE FROM wasde_data WHERE commodity = ? "
            f"AND statistic_category IN ({','.join(['?'] * len(cats))}) "
            f"AND agg_level IN ('NATIONAL', 'STATE')",
            (commodity, *cats))
        deleted = cur.rowcount
        rows = []
        n_file = 0
        for block in payload["blocks"].values():
            for r in block["data"]:
                rows.append((
                    r.get("commodity_desc", commodity).upper(),
                    r.get("statisticcat_desc", ""),
                    r.get("Value", r.get("value")),
                    numeric(r.get("Value", r.get("value"))),
                    r.get("unit_desc", ""),
                    r.get("location_desc", r.get("state_name", "Unknown")),
                    r.get("state_alpha", ""),
                    r.get("agg_level_desc", ""),
                    int(r["year"]) if r.get("year") else None,
                    r.get("reference_period_desc", ""),
                    r.get("short_desc", ""),
                    r.get("source_desc", ""),
                    r.get("sector_desc", ""),
                    r.get("group_desc", ""),
                    r.get("class_desc", ""),
                    r.get("freq_desc", ""),
                    r.get("load_time"),
                    NOW,
                    r.get("prodn_practice_desc", ""),
                    r.get("util_practice_desc", ""),
                    r.get("domain_desc", ""),
                    r.get("domaincat_desc", ""),
                ))
                if len(rows) >= BATCH:
                    cur.executemany(ins, rows)
                    n_file += len(rows)
                    rows = []
        if rows:
            cur.executemany(ins, rows)
            n_file += len(rows)
        con.commit()
        total += n_file
        log(f"NASS {commodity}: -{deleted} snapshot rows, +{n_file} history rows")
    return total


# ---------------------------------------------------------------------------
# 2. FAOSTAT producer prices (USD/tonne, annual) -> global_prices
# ---------------------------------------------------------------------------

def import_fao_producer(con: sqlite3.Connection) -> int:
    if not FAO_PRICES_CSV.exists():
        log("FAOSTAT prices CSV missing — SKIPPING")
        return 0
    cur = con.cursor()
    cur.execute("DELETE FROM global_prices WHERE source = 'FAOSTAT'")
    log(f"FAOSTAT producer: cleared {cur.rowcount} prior rows")
    ins = ("INSERT INTO global_prices (commodity, price, currency, unit, region, "
           "country, date, source, indicator_code, imported_at) "
           "VALUES (?,?,?,?,?,?,?,?,?,?)")
    rows, total = [], 0
    with FAO_PRICES_CSV.open(encoding="utf-8-sig", newline="") as f:
        for r in csv.DictReader(f):
            if r["Element"] != "Producer Price (USD/tonne)":
                continue
            if r["Months"] != "Annual value":
                continue
            v = numeric(r["Value"])
            if v is None:
                continue
            rows.append((
                r["Item"], v, "USD", "USD/tonne", r["Area"], r["Area"],
                f"{r['Year']}-07-01 00:00:00.000000",
                "FAOSTAT", f"PP_{r['Item Code']}", NOW))
            if len(rows) >= BATCH:
                cur.executemany(ins, rows)
                total += len(rows)
                rows = []
    if rows:
        cur.executemany(ins, rows)
        total += len(rows)
    con.commit()
    log(f"FAOSTAT producer prices: +{total} rows")
    return total


# ---------------------------------------------------------------------------
# 3. FAOSTAT food CPI (monthly per country) -> global_prices
# ---------------------------------------------------------------------------

def import_fao_cpi(con: sqlite3.Connection) -> int:
    if not FAO_CPI_CSV.exists():
        log("FAOSTAT CPI CSV missing — SKIPPING")
        return 0
    cur = con.cursor()
    cur.execute("DELETE FROM global_prices WHERE source = 'FAOSTAT CPI'")
    log(f"FAOSTAT CPI: cleared {cur.rowcount} prior rows")
    ins = ("INSERT INTO global_prices (commodity, price, currency, unit, region, "
           "country, date, source, indicator_code, imported_at) "
           "VALUES (?,?,?,?,?,?,?,?,?,?)")
    rows, total = [], 0
    with FAO_CPI_CSV.open(encoding="utf-8-sig", newline="") as f:
        for r in csv.DictReader(f):
            if r["Item Code"] != "23013":
                continue
            month = MONTHS.get(r["Months"])
            if month is None:
                continue
            v = numeric(r["Value"])
            if v is None:
                continue
            rows.append((
                "Consumer Food Price Index (2015=100)", v, "", "index 2015=100",
                r["Area"], r["Area"],
                f"{r['Year']}-{month:02d}-01 00:00:00.000000",
                "FAOSTAT CPI", "FAO_CPI_FOOD", NOW))
            if len(rows) >= BATCH:
                cur.executemany(ins, rows)
                total += len(rows)
                rows = []
    if rows:
        cur.executemany(ins, rows)
        total += len(rows)
    con.commit()
    log(f"FAOSTAT food CPI: +{total} rows")
    return total


# ---------------------------------------------------------------------------
# 4. World Bank Pink Sheet (monthly) -> global_prices
# ---------------------------------------------------------------------------

def import_pinksheet(con: sqlite3.Connection) -> int:
    if not PINKSHEET_XLSX.exists():
        log("Pink Sheet xlsx missing — SKIPPING")
        return 0
    try:
        from openpyxl import load_workbook
    except ImportError:
        log("openpyxl not installed — run: pip install openpyxl")
        return 0
    cur = con.cursor()
    cur.execute("DELETE FROM global_prices WHERE source = 'World Bank Pink Sheet'")
    log(f"Pink Sheet: cleared {cur.rowcount} prior rows")

    wb = load_workbook(PINKSHEET_XLSX, read_only=True, data_only=True)
    ws = wb["Monthly Prices"]
    it = ws.iter_rows(values_only=True)
    rows_raw = list(it)
    # Layout: row with commodity names, then a units row, then data rows
    # whose first cell looks like 1960M01. Find the header rows dynamically.
    name_row = units_row = None
    data_start = None
    for i, row in enumerate(rows_raw[:10]):
        first = str(row[0]).strip() if row[0] is not None else ""
        if first and len(first) == 7 and first[4] == "M" and first[:4].isdigit():
            data_start = i
            break
    if data_start is None:
        log("Pink Sheet: could not find data rows — SKIPPING")
        return 0
    # names = last row before data that has >5 non-null string cells;
    # units row sits directly above the data.
    for i in range(data_start - 1, -1, -1):
        cells = [c for c in rows_raw[i][1:] if isinstance(c, str) and c.strip()]
        if len(cells) > 5:
            if units_row is None:
                units_row = rows_raw[i]
            else:
                name_row = rows_raw[i]
                break
    if name_row is None:
        name_row, units_row = units_row, None
    names = ["" if c is None else str(c).strip() for c in name_row]
    units = (["" if c is None else str(c).strip() for c in units_row]
             if units_row else [""] * len(names))

    ins = ("INSERT INTO global_prices (commodity, price, currency, unit, region, "
           "country, date, source, indicator_code, imported_at) "
           "VALUES (?,?,?,?,?,?,?,?,?,?)")
    batch, total = [], 0
    for row in rows_raw[data_start:]:
        first = str(row[0]).strip() if row[0] is not None else ""
        if len(first) != 7 or first[4] != "M":
            continue
        year, month = int(first[:4]), int(first[5:7])
        date = f"{year}-{month:02d}-01 00:00:00.000000"
        for ci in range(1, len(row)):
            v = row[ci]
            if v is None or not isinstance(v, (int, float)):
                continue
            name = names[ci] if ci < len(names) else ""
            if not name:
                continue
            unit = units[ci] if ci < len(units) else ""
            batch.append((name, float(v), "USD", unit, "World", "World",
                          date, "World Bank Pink Sheet", f"CMO_{ci}", NOW))
            if len(batch) >= BATCH:
                cur.executemany(ins, batch)
                total += len(batch)
                batch = []
    if batch:
        cur.executemany(ins, batch)
        total += len(batch)
    con.commit()
    log(f"Pink Sheet: +{total} rows ({len([n for n in names if n])} series)")
    return total


# ---------------------------------------------------------------------------
# 5. BLS AP retail series -> retail_prices
# ---------------------------------------------------------------------------

def import_bls_ap(con: sqlite3.Connection) -> int:
    files = sorted(BLS_AP_DIR.glob("ap_fred_*.json"))
    if not files:
        log("BLS AP: no files — SKIPPING")
        return 0
    cur = con.cursor()
    cur.execute("DELETE FROM retail_prices WHERE source = 'BLS AP'")
    log(f"BLS AP: cleared {cur.rowcount} prior rows")
    ins = ("INSERT INTO retail_prices (food_item, price, unit, store_type, "
           "location, state, country, date, source, brand, quality_grade, "
           "imported_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)")
    total = 0
    for fp in files:
        p = json.loads(fp.read_text(encoding="utf-8"))
        # `location` is the publisher's own area, taken from the artifact — NOT
        # hardcoded. Before 2026-07-24 every row was stamped 'U.S. city average'
        # because the store held only area 0000 series; the four Census-region
        # tomato series (APU0[1-4]00712311) would then have been silently
        # mislabelled as national. Older artifacts have no `area` key, and their
        # area really is U.S. city average, so that stays the default.
        area = p.get("area") or "U.S. city average"
        rows = [(
            p["item_name"], r["value"], p["unit"], "Grocery (avg)",
            area, "", "USA",
            f"{r['date']} 00:00:00.000000", "BLS AP", "", "", NOW)
            for r in p["data"]]
        cur.executemany(ins, rows)
        total += len(rows)
    con.commit()
    log(f"BLS AP: +{total} rows across {len(files)} items")
    return total


# ---------------------------------------------------------------------------
# 6. FRED / BLS macro + food CPI/PPI series -> economic_indicators   (P0-5)
# ---------------------------------------------------------------------------

def _indicator_category(series_id: str) -> str:
    """Classify a series id into the category vocabulary already in the table."""
    if series_id.startswith("WPU"):
        return "PPI"
    if series_id.startswith(("CUUR", "CUSR")) and ("SAF" in series_id or "SEF" in series_id):
        return "Food CPI"
    if series_id.startswith("CPI"):
        return "CPI"
    return "Macro"


def import_economic_indicators(con: sqlite3.Connection) -> int:
    """
    Refresh `economic_indicators` from the canonical collected artifacts.

    P0-5: before 2026-07-24 the rebake only COUNTED this table (and
    `composite_indices`); neither was ever refreshed, so both froze at their
    last ad-hoc load while every other table moved on.

    Upsert on (series_id, date, source): values and `imported_at` are
    refreshed in place, new observations are inserted, and nothing is deleted -
    so series loaded by other collectors survive a rebake untouched.
    """
    total_seen = updated = inserted = dupe_groups = 0
    cur = con.cursor()

    for path, source in COLLECTED_SOURCES:
        if not path.exists():
            log(f"economic_indicators: {path.name} missing — SKIPPING {source}")
            continue

        payload = json.loads(path.read_text(encoding="utf-8"))
        n_series = 0
        for series_id, block in payload.items():
            observations = block.get("data") or []
            if not observations:
                continue
            n_series += 1
            name = block.get("name") or series_id
            category = _indicator_category(series_id)

            for obs in observations:
                date_str = obs.get("date")
                value = numeric(obs.get("value"))
                if not date_str or value is None:
                    continue
                total_seen += 1
                date_full = f"{date_str} 00:00:00.000000"

                rows = cur.execute(
                    "SELECT id FROM economic_indicators "
                    "WHERE series_id = ? AND date = ? AND source = ?",
                    (series_id, date_full, source)).fetchall()

                if rows:
                    if len(rows) > 1:
                        dupe_groups += 1
                    cur.executemany(
                        "UPDATE economic_indicators SET value = ?, imported_at = ? "
                        "WHERE id = ?",
                        [(value, NOW, r[0]) for r in rows])
                    updated += len(rows)
                else:
                    cur.execute(
                        "INSERT INTO economic_indicators (indicator_name, series_id, "
                        "value, date, category, frequency, source, imported_at) "
                        "VALUES (?,?,?,?,?,?,?,?)",
                        (name, series_id, value, date_full, category,
                         "Monthly", source, NOW))
                    inserted += 1

        log(f"economic_indicators: {source} — {n_series} series from {path.name}")

    con.commit()
    if dupe_groups:
        log(f"economic_indicators: WARNING — {dupe_groups} (series_id,date,source) "
            f"groups hold duplicate rows; all copies were refreshed, none removed")
    log(f"economic_indicators: {total_seen} observations processed "
        f"(+{inserted} inserted, {updated} rows updated)")
    return inserted + updated


# ---------------------------------------------------------------------------
# 7. composite_indices — recomputed from the refreshed source tables  (P0-5)
# ---------------------------------------------------------------------------

def refresh_composite_indices(con: sqlite3.Connection) -> int:
    """
    Rebuild `composite_indices` from the *current* contents of `global_prices`
    (FAO) and `economic_indicators` (BLS CPI).

    P0-5: this step did not exist. `composite_indices.computed_at` was
    uniformly 2026-03-28, so the headline Food Price Index page served whatever
    was computed that day regardless of later source updates.

    Must run AFTER import_economic_indicators, since `bls_overall` is derived
    from it. Writes are upserts (see indices/composite.py), so this is
    idempotent and does actually update changed values.
    """
    from indices.composite import compute_all_indices

    stats = compute_all_indices(conn=con)
    con.commit()
    log(f"composite_indices: +{stats['inserted']} inserted, "
        f"{stats['updated']} updated, {stats['unchanged']} unchanged")
    return stats["inserted"] + stats["updated"]


def main() -> int:
    if not DB_PATH.exists():
        print(f"DB not found: {DB_PATH}")
        return 1
    size0 = DB_PATH.stat().st_size / 1048576
    con = sqlite3.connect(str(DB_PATH))
    con.execute("PRAGMA journal_mode=WAL")
    con.execute("PRAGMA synchronous=NORMAL")

    counts = {
        "nass": import_nass(con),
        "fao_producer": import_fao_producer(con),
        "fao_cpi": import_fao_cpi(con),
        "pinksheet": import_pinksheet(con),
        "bls_ap": import_bls_ap(con),
        # P0-5: these two were previously only counted, never refreshed.
        # economic_indicators must precede composites (bls_overall reads it).
        "economic_indicators": import_economic_indicators(con),
        "composite_indices": refresh_composite_indices(con),
    }

    # Helpful composite indexes for the new query patterns.
    con.execute("CREATE INDEX IF NOT EXISTS ix_global_source_commodity "
                "ON global_prices (source, commodity, date)")
    con.execute("CREATE INDEX IF NOT EXISTS ix_global_country "
                "ON global_prices (country, commodity)")
    con.execute("CREATE INDEX IF NOT EXISTS ix_retail_source_item "
                "ON retail_prices (source, food_item, date)")
    con.commit()

    # Sync bookkeeping rows.
    for name, n in counts.items():
        con.execute(
            "INSERT INTO data_source_sync (source_name, last_sync_time, "
            "last_sync_status, records_synced, sync_frequency) VALUES (?,?,?,?,?) "
            "ON CONFLICT(source_name) DO UPDATE SET last_sync_time=excluded.last_sync_time, "
            "last_sync_status=excluded.last_sync_status, records_synced=excluded.records_synced",
            (f"rebake_{name}", NOW, "SUCCESS", n, "manual"))
    con.commit()

    log("VACUUM (this can take a minute)…")
    con.execute("PRAGMA journal_mode=DELETE")
    con.execute("VACUUM")
    con.close()

    size1 = DB_PATH.stat().st_size / 1048576
    log(f"counts: {counts}")
    log(f"DB size: {size0:.0f} MB -> {size1:.0f} MB")

    # Final table census.
    con = sqlite3.connect(str(DB_PATH))
    for t in ("wasde_data", "global_prices", "retail_prices",
              "economic_indicators", "composite_indices"):
        n = con.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
        log(f"  {t}: {n:,} rows")
    con.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
