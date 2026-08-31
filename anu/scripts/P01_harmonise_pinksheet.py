#!/usr/bin/env python3
"""P01 — Harmonise the World Bank Pink Sheet into tidy per-series CSVs.

Input:  data/raw/pinksheet/CMO-Historical-Data-Monthly.xlsx ("Monthly Prices")
Output: data/final/pink_sheet/ps_<slug>.csv  — one file per commodity column,
        columns: date (YYYY-MM-01), value, unit (the publisher's own token).

The wide sheet is read exactly as published: names row, units row, then data
rows keyed by YYYYMMM period labels. Values are verbatim; cells the publisher
left blank or '..' are skipped, never imputed.
"""

import csv
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from _shared import RAW, FINAL, slugify  # noqa: E402

SRC = RAW / "pinksheet" / "CMO-Historical-Data-Monthly.xlsx"
DEST_DIR = FINAL / "pink_sheet"


def main() -> int:
    if not SRC.exists():
        print(f"[P01] missing input {SRC} — run L01 first", file=sys.stderr)
        return 1
    from openpyxl import load_workbook
    wb = load_workbook(SRC, read_only=True, data_only=True)
    ws = wb["Monthly Prices"]
    rows = list(ws.iter_rows(values_only=True))

    # Locate the first data row (period label like 1960M01).
    data_start = None
    for i, row in enumerate(rows[:12]):
        first = str(row[0]).strip() if row[0] is not None else ""
        if len(first) == 7 and first[4] == "M" and first[:4].isdigit():
            data_start = i
            break
    if data_start is None:
        print("[P01] FAIL: could not locate data rows", file=sys.stderr)
        return 1

    # names row = last header row with >5 non-empty string cells above the
    # units row; units row = the header row directly above the data.
    name_row = units_row = None
    for i in range(data_start - 1, -1, -1):
        cells = [c for c in rows[i][1:] if isinstance(c, str) and c.strip()]
        if len(cells) > 5:
            if units_row is None:
                units_row = rows[i]
            else:
                name_row = rows[i]
                break
    if name_row is None:
        name_row, units_row = units_row, None
    names = ["" if c is None else str(c).strip() for c in name_row]
    units = ["" if c is None else str(c).strip() for c in units_row] \
        if units_row else [""] * len(names)

    DEST_DIR.mkdir(parents=True, exist_ok=True)
    written = 0
    for ci in range(1, len(names)):
        name = names[ci]
        if not name:
            continue
        clean = name.replace("**", "").strip()
        sid = f"ps_{slugify(clean)}"
        unit = units[ci] if ci < len(units) else ""
        out_rows = []
        for row in rows[data_start:]:
            first = str(row[0]).strip() if row[0] is not None else ""
            if len(first) != 7 or first[4] != "M":
                continue
            v = row[ci] if ci < len(row) else None
            if not isinstance(v, (int, float)):
                continue
            out_rows.append((f"{first[:4]}-{first[5:7]}-01", float(v), unit))
        if not out_rows:
            continue
        dest = DEST_DIR / f"{sid}.csv"
        with open(dest, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["date", "value", "unit"])
            w.writerows(out_rows)
        written += 1
        print(f"[P01] {sid}: {len(out_rows)} obs "
              f"({out_rows[0][0]}..{out_rows[-1][0]}) [{clean} {unit}]")
    print(f"[P01] wrote {written} series to {DEST_DIR}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
