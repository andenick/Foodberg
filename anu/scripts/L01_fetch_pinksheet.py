#!/usr/bin/env python3
"""L01 — Fetch the World Bank Pink Sheet (monthly commodity prices).

Source (public, keyless, CC BY 4.0):
    https://www.worldbank.org/en/research/commodity-markets
    -> CMO-Historical-Data-Monthly.xlsx  ("Monthly Prices" sheet)

Strategy:
  1. Scrape the commodity-markets landing page for the current
     CMO-Historical-Data-Monthly.xlsx href (the per-edition document hash
     rotates; this is the ONLY way to guarantee the current edition — the
     long-lived "stable" URL is frozen at a stale 2024-12 edition).
  2. If the scrape fails, fall back to the pinned document URL (verified live
     2026-08; serves the frozen Dec-2024 vintage — usable, but not current).

Output:
    data/raw/pinksheet/CMO-Historical-Data-Monthly.xlsx
    data/raw/pinksheet/CMO-Historical-Data-Monthly.xlsx.fetch_meta.json
"""

import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from _shared import RAW, http_get, write_fetch_meta, UA  # noqa: E402

PINNED_URL = (
    "https://thedocs.worldbank.org/en/doc/"
    "5d903e848db1d1b83e0ec8f744e55570-0350012021/related/"
    "CMO-Historical-Data-Monthly.xlsx"
)
LANDING = "https://www.worldbank.org/en/research/commodity-markets"
DEST = RAW / "pinksheet" / "CMO-Historical-Data-Monthly.xlsx"
MIN_BYTES = 100_000  # a real edition is ~750 KB


def scrape_current_url() -> str:
    import urllib.request
    req = urllib.request.Request(LANDING, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=60) as r:
        html = r.read().decode("utf-8", errors="replace")
    hits = re.findall(r'href="([^"]*CMO-Historical-Data-Monthly\.xlsx[^"]*)"', html)
    if not hits:
        raise RuntimeError("No Pink Sheet XLSX link found on the landing page")
    url = hits[0]
    if url.startswith("/"):
        url = "https://www.worldbank.org" + url
    return url


def edition_date(dest: Path) -> str:
    """The publisher's own 'Updated on ...' banner line, if present."""
    from openpyxl import load_workbook
    try:
        wb = load_workbook(dest, read_only=True, data_only=True)
        ws = wb["Monthly Prices"]
        for row in ws.iter_rows(min_row=1, max_row=8, max_col=1, values_only=True):
            first = str(row[0]).strip() if row[0] is not None else ""
            if first.lower().startswith("updated on"):
                return first
    except Exception:  # noqa: BLE001 — banner is informational only
        pass
    return "unknown"


def main() -> int:
    url_used = None
    try:
        url_used = scrape_current_url()
        print(f"[L01] current edition URL: {url_used}")
        http_get(url_used, DEST)
    except Exception as e:  # noqa: BLE001
        print(f"[L01] landing-page scrape failed ({e}); falling back to pinned URL")
        url_used = PINNED_URL
        http_get(url_used, DEST)

    size = DEST.stat().st_size
    if size < MIN_BYTES:
        print(f"[L01] FAIL: suspiciously small file ({size} bytes)", file=sys.stderr)
        return 1
    write_fetch_meta(DEST, url_used, edition=edition_date(DEST))
    print(f"[L01] Pink Sheet saved: {DEST} ({size:,} bytes) "
          f"[{edition_date(DEST)}] from {url_used}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
